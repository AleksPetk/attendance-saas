"""Excel (.xlsx) exporter for Attendance Report payloads."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from attendance.report_export.common import (
    cell_text,
    flatten_report_rows,
    period_lines,
    report_identity_headers,
    report_identity_values,
    report_shows_class_column,
    status_label,
)


def build_attendance_report_xlsx(report: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Report"

    bold = Font(bold=True, name="Calibri", size=12)
    title_font = Font(bold=True, name="Calibri", size=14)
    meta_font = Font(name="Calibri", size=11, color="445566")
    header_font = Font(bold=True, name="Calibri", size=11, color="1F2937")
    header_fill = PatternFill("solid", fgColor="EEF2F7")

    row_idx = 1
    sheet.cell(row_idx, 1, report.get("organization_name") or "").font = meta_font
    row_idx += 1

    sheet.cell(row_idx, 1, report.get("group_name") or "Group").font = title_font
    row_idx += 1

    status = status_label(report.get("group_status"))
    if status:
        sheet.cell(row_idx, 1, f"Status: {status}").font = meta_font
        row_idx += 1

    sheet.cell(row_idx, 1, "Attendance Report").font = bold
    row_idx += 1

    for line in period_lines(report):
        sheet.cell(row_idx, 1, line).font = meta_font
        row_idx += 1

    row_idx += 1
    columns = list(report.get("columns") or [])
    headers = [
        *report_identity_headers(report),
        *[col.get("label") or col.get("key") for col in columns],
    ]
    header_row = row_idx
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    sheet.freeze_panes = f"A{header_row + 1}"

    identity_count = 3 if report_shows_class_column(report) else 2
    for row in flatten_report_rows(report):
        row_idx += 1
        values = [
            *report_identity_values(row, report),
            *[cell_text(row["cells"].get(col["key"])) for col in columns],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_idx > identity_count:
                cell.number_format = "@"

    if report_shows_class_column(report):
        widths = [18, 16, 22] + [14] * len(columns)
    else:
        widths = [18, 22] + [14] * len(columns)
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
