import base64
import datetime
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from reportlab.pdfbase.pdfmetrics import stringWidth
from rest_framework.test import APIClient

from attendance.models import ActionRecord, ActionSource, ActionType
from attendance.report_export import (
    build_attendance_report_csv,
    build_attendance_report_pdf,
    build_attendance_report_xlsx,
    build_export_filename,
    enrich_report_for_export,
)
from attendance.report_export.fonts import font_for_text
from groups.deletion import permanently_delete_group
from groups.models import Group, GroupStatus
from members.models import Member
from organizations.models import Organization

User = get_user_model()


def create_user(email, *, password="secure-password", verified=True, **extra_fields):
    user = User.objects.create_user(email=email, password=password, **extra_fields)
    if verified:
        user.mark_email_verified()
    return user


def basic_auth_header(identity, password):
    token = base64.b64encode(f"{identity}:{password}".encode()).decode()
    return f"Basic {token}"


SAMPLE_REPORT = {
    "group_name": "SELS Kids",
    "group_status": "active",
    "source_group_id": 12,
    "date_preset": "custom",
    "date_from": "2026-12-12",
    "date_to": "2026-12-25",
    "date_label": "12 December 2026 - 25 December 2026",
    "organization_name": "Demo Org",
    "columns": [
        {"key": "check_in", "label": "Check-in"},
        {"key": "break", "label": "Break"},
        {"key": "check_out", "label": "Check-out"},
    ],
    "sections": [
        {
            "date": "2026-12-15",
            "label": "15 December 2026",
            "rows": [
                {
                    "participant_key": "member:1",
                    "name": "Aleks",
                    "cells": {"check_in": "09:00", "break": "12:00", "check_out": "17:30"},
                },
                {
                    "participant_key": "member:2",
                    "name": "Нами",
                    "cells": {"check_in": "09:05", "break": None, "check_out": "17:25"},
                },
                {
                    "participant_key": "member:3",
                    "name": "田中",
                    "cells": {"check_in": "09:10", "break": "12:15", "check_out": "17:40"},
                },
            ],
        }
    ],
}


class AttendanceReportExportUnitTests(TestCase):
    def test_csv_utf8_and_content(self):
        csv_text = build_attendance_report_csv(SAMPLE_REPORT)
        self.assertIn("SELS Kids", csv_text)
        self.assertIn("Attendance Report", csv_text)
        self.assertIn("Date,Name,Check-in,Break,Check-out", csv_text)
        self.assertIn("Aleks,09:00,12:00,17:30", csv_text)
        self.assertIn("Нами,09:05,,17:25", csv_text)
        self.assertIn("田中,09:10,12:15,17:40", csv_text)

    def test_csv_includes_archived_status(self):
        report = {**SAMPLE_REPORT, "group_status": "archived"}
        csv_text = build_attendance_report_csv(report)
        self.assertIn("Archived", csv_text)

    def test_excel_columns_and_unicode(self):
        content = build_attendance_report_xlsx(SAMPLE_REPORT)
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active
        values = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=1, max_col=5)]
        flat = [v for row in values for v in row if v]
        self.assertIn("SELS Kids", flat)
        self.assertIn("Attendance Report", flat)
        self.assertIn("Date", flat)
        self.assertIn("Aleks", flat)
        self.assertIn("Нами", flat)
        self.assertIn("田中", flat)
        self.assertIn("09:00", flat)

    def test_pdf_generation_and_unicode_fonts(self):
        pdf_bytes = build_attendance_report_pdf(SAMPLE_REPORT)
        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertGreater(len(pdf_bytes), 1000)
        # Font selection must support Cyrillic and CJK.
        self.assertEqual(font_for_text("Нами"), "AttendanceReportSans")
        self.assertEqual(font_for_text("田中"), "HeiseiKakuGo-W5")
        # Glyph metrics should be available for both font paths.
        self.assertGreater(stringWidth("Нами", "AttendanceReportSans", 10), 0)
        self.assertGreater(stringWidth("田中", "HeiseiKakuGo-W5", 10), 0)

    def test_filename_rules(self):
        self.assertEqual(
            build_export_filename(SAMPLE_REPORT, "xlsx"),
            "sels-kids-attendance-2026-12-12-to-2026-12-25.xlsx",
        )
        month_report = {
            **SAMPLE_REPORT,
            "date_from": "2026-12-01",
            "date_to": "2026-12-31",
            "date_label": "December 2026",
        }
        self.assertEqual(
            build_export_filename(month_report, "pdf"),
            "sels-kids-attendance-december-2026.pdf",
        )

    def test_enrich_organization_name(self):
        org = Organization(workspace_id="WS123456", internal_label="North Branch")
        enriched = enrich_report_for_export({"group_name": "G"}, organization=org)
        self.assertEqual(enriched["organization_name"], "North Branch")


class AttendanceReportExportApiTests(TestCase):
    def setUp(self):
        self.owner = create_user("export-owner@example.com")
        self.org = Organization.objects.create_with_owner(
            owner=self.owner, internal_label="Export Org"
        )
        self.other_owner = create_user("export-other@example.com")
        self.other_org = Organization.objects.create_with_owner(owner=self.other_owner)
        self.client = APIClient()
        self.client.credentials(
            HTTP_AUTHORIZATION=basic_auth_header(self.owner.email, "secure-password")
        )
        self.other_client = APIClient()
        self.other_client.credentials(
            HTTP_AUTHORIZATION=basic_auth_header(self.other_owner.email, "secure-password")
        )
        self.group = Group.objects.create_group(organization=self.org, name="SELS Kids")
        self.member = Member.objects.create(
            organization=self.org,
            name="Aleks",
            email="aleks-export@example.com",
        )
        now = timezone.now().replace(hour=9, minute=0, second=0, microsecond=0)
        ActionRecord.objects.create(
            organization=self.org,
            group=self.group,
            source_group_id=self.group.pk,
            participant_kind="member",
            member=self.member,
            action_type=ActionType.CHECK_IN,
            source=ActionSource.KIOSK,
            performed_at=now,
            participant_name_snapshot="Aleks",
            group_name_snapshot=self.group.name,
        )

    def _export(self, client, **params):
        query = {
            "source_group_id": self.group.pk,
            "preset": "today",
            "export_format": "csv",
            **params,
        }
        return client.get("/api/history/attendance-report/export/", query)

    def test_export_csv_xlsx_pdf(self):
        for fmt, mime_part in (
            ("csv", "text/csv"),
            ("xlsx", "spreadsheetml"),
            ("pdf", "pdf"),
        ):
            resp = self._export(self.client, export_format=fmt)
            self.assertEqual(resp.status_code, 200, fmt)
            self.assertIn(mime_part, resp["Content-Type"], fmt)
            self.assertIn("attachment;", resp["Content-Disposition"], fmt)
            self.assertTrue(len(resp.content) > 20, fmt)

        csv_resp = self._export(self.client, export_format="csv")
        # UTF-8 BOM for Excel compatibility
        self.assertTrue(csv_resp.content.startswith(b"\xef\xbb\xbf"))
        body = csv_resp.content.decode("utf-8-sig")
        self.assertIn("SELS Kids", body)
        self.assertIn("Aleks", body)

    def test_custom_range_export(self):
        day = timezone.now().date()
        resp = self._export(
            self.client,
            export_format="csv",
            preset="custom",
            date_from=day.isoformat(),
            date_to=day.isoformat(),
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn("Aleks", resp.content.decode("utf-8-sig"))

    def test_deleted_group_export(self):
        self.group.status = GroupStatus.ARCHIVED
        self.group.save(update_fields=["status", "archived_at", "updated_at"])
        group_id = self.group.pk
        permanently_delete_group(self.group)
        resp = self.client.get(
            "/api/history/attendance-report/export/",
            {
                "source_group_id": group_id,
                "preset": "today",
                "export_format": "csv",
            },
        )
        self.assertEqual(resp.status_code, 200)
        text = resp.content.decode("utf-8-sig")
        self.assertIn("Deleted", text)
        self.assertIn("SELS Kids", text)

    def test_tenant_isolation(self):
        resp = self._export(self.other_client, export_format="pdf")
        self.assertEqual(resp.status_code, 404)

    def test_empty_report_rejected(self):
        empty_day = (timezone.now() - datetime.timedelta(days=40)).date()
        resp = self._export(
            self.client,
            export_format="csv",
            preset="custom",
            date_from=empty_day.isoformat(),
            date_to=empty_day.isoformat(),
        )
        self.assertEqual(resp.status_code, 400)
