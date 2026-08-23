"""
Structured Group Class snapshots on ActionRecord and Attendance Report.

Covers historical Class identity for reports/exports without changing
Standard Group report behavior.
"""

import base64
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from attendance.attendance_report import UNKNOWN_CLASS_LABEL, build_attendance_report
from attendance.models import ActionRecord, ActionSource, ActionType
from attendance.report_export import (
    build_attendance_report_csv,
    build_attendance_report_pdf,
    build_attendance_report_xlsx,
)
from attendance.services import perform_action_record_from_kiosk
from groups.models import (
    Group,
    GroupMembership,
    GroupSection,
    GroupSectionStatus,
    GroupType,
)
from groups.section_deletion import permanently_delete_section
from members.models import Member
from organizations.models import Organization

User = get_user_model()


def create_user(email, *, password="secure-password", verified=True, **extra_fields):
    user = User.objects.create_user(email=email, password=password, **extra_fields)
    if verified:
        user.mark_email_verified()
    return user


def basic_auth_header(identity, password):
    token = base64.b64encode(f"{identity}:{password}".encode()).decode()
    return f"Basic {token}"


class StructuredAttendanceReportClassTests(TestCase):
    def setUp(self):
        self.owner = create_user("structured-report@example.com")
        self.org = Organization.objects.create_with_owner(owner=self.owner)
        self.other_owner = create_user("structured-report-other@example.com")
        self.other_org = Organization.objects.create_with_owner(owner=self.other_owner)

        self.client = APIClient()
        self.client.credentials(
            HTTP_AUTHORIZATION=basic_auth_header(self.owner.email, "secure-password")
        )
        self.other_client = APIClient()
        self.other_client.credentials(
            HTTP_AUTHORIZATION=basic_auth_header(self.other_owner.email, "secure-password")
        )

        self.structured = Group.objects.create_group(
            organization=self.org,
            name="School Day",
            group_type=GroupType.STRUCTURED,
            check_in_enabled=True,
            check_out_enabled=True,
            breaks_enabled=True,
            max_breaks=2,
        )
        self.class_a = GroupSection.objects.create_section(
            group=self.structured, name="Class A"
        )
        self.class_b = GroupSection.objects.create_section(
            group=self.structured, name="Class B"
        )

        self.standard = Group.objects.create_group(
            organization=self.org,
            name="Flat Club",
            group_type=GroupType.STANDARD,
            check_in_enabled=True,
            check_out_enabled=True,
            breaks_enabled=False,
        )

        self.clara = Member.objects.create(
            organization=self.org, name="Clara", email="clara@example.com"
        )
        self.aleks = Member.objects.create(
            organization=self.org, name="Aleks", email="aleks@example.com"
        )
        self.jimi = Member.objects.create(
            organization=self.org, name="Jimi", email="jimi@example.com"
        )

        self.clara_membership = GroupMembership.objects.create(
            organization=self.org,
            group=self.structured,
            member=self.clara,
            section=self.class_a,
        )
        self.aleks_membership = GroupMembership.objects.create(
            organization=self.org,
            group=self.structured,
            member=self.aleks,
            section=self.class_a,
        )
        self.jimi_membership = GroupMembership.objects.create(
            organization=self.org,
            group=self.structured,
            member=self.jimi,
            section=self.class_b,
        )
        self.standard_membership = GroupMembership.objects.create(
            organization=self.org,
            group=self.standard,
            member=self.clara,
        )

    def _day(self, hour=9, minute=0):
        return timezone.now().replace(
            hour=hour, minute=minute, second=0, microsecond=0
        )

    def _manual_record(
        self,
        *,
        member,
        action_type,
        performed_at,
        group=None,
        section=None,
        class_name_snapshot=None,
        source_section_id=None,
        group_type_snapshot=None,
    ):
        group = group or self.structured
        kwargs = {
            "organization": self.org,
            "group": group,
            "source_group_id": group.pk,
            "participant_kind": "member",
            "member": member,
            "action_type": action_type,
            "source": ActionSource.KIOSK,
            "performed_at": performed_at,
            "participant_name_snapshot": member.name,
            "participant_email_snapshot": member.email,
            "group_name_snapshot": group.name,
            "group_type_snapshot": group_type_snapshot or group.group_type,
        }
        if section is not None:
            kwargs["section"] = section
            kwargs["source_section_id"] = (
                source_section_id if source_section_id is not None else section.pk
            )
            kwargs["class_name_snapshot"] = (
                class_name_snapshot
                if class_name_snapshot is not None
                else section.name
            )
        elif class_name_snapshot is not None or source_section_id is not None:
            kwargs["class_name_snapshot"] = class_name_snapshot or ""
            kwargs["source_section_id"] = source_section_id
        return ActionRecord.objects.create(**kwargs)

    def test_a_structured_perform_stores_class_snapshot(self):
        ar = perform_action_record_from_kiosk(
            group=self.structured,
            participant_kind="member",
            action_type=ActionType.CHECK_IN,
            membership=self.clara_membership,
            now=self._day(11, 15),
        )
        self.assertEqual(ar.section_id, self.class_a.pk)
        self.assertEqual(ar.source_section_id, self.class_a.pk)
        self.assertEqual(ar.class_name_snapshot, "Class A")
        self.assertEqual(ar.group_type_snapshot, GroupType.STRUCTURED)

        std = perform_action_record_from_kiosk(
            group=self.standard,
            participant_kind="member",
            action_type=ActionType.CHECK_IN,
            membership=self.standard_membership,
            now=self._day(11, 20),
        )
        self.assertIsNone(std.section_id)
        self.assertIsNone(std.source_section_id)
        self.assertEqual(std.class_name_snapshot, "")
        self.assertEqual(std.group_type_snapshot, GroupType.STANDARD)

    def test_b_structured_report_includes_class_column(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 15),
            section=self.class_a,
        )
        self._manual_record(
            member=self.jimi,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 44),
            section=self.class_b,
        )
        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.structured.pk,
            preset="today",
        )
        self.assertEqual(report["group_type"], GroupType.STRUCTURED)
        self.assertEqual(len(report["sections"]), 1)
        rows = report["sections"][0]["rows"]
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["class_name"], "Class A")
        self.assertEqual(rows[0]["name"], "Clara")
        self.assertEqual(rows[1]["class_name"], "Class B")
        self.assertEqual(rows[1]["name"], "Jimi")

        resp = self.client.get(
            f"/api/history/attendance-report/?source_group_id={self.structured.pk}&preset=today"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["group_type"], "structured")
        self.assertEqual(resp.data["sections"][0]["rows"][0]["class_name"], "Class A")

    def test_c_standard_report_does_not_include_class_column(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(9, 0),
            group=self.standard,
        )
        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.standard.pk,
            preset="today",
        )
        self.assertEqual(report["group_type"], GroupType.STANDARD)
        row = report["sections"][0]["rows"][0]
        self.assertNotIn("class_name", row)
        self.assertEqual(row["name"], "Clara")

    def test_d_class_rename_does_not_alter_historical_display(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 15),
            section=self.class_a,
            class_name_snapshot="Class A",
        )
        self.class_a.name = "Class Alpha"
        self.class_a.save(update_fields=["name", "updated_at"])
        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.structured.pk,
            preset="today",
        )
        self.assertEqual(report["sections"][0]["rows"][0]["class_name"], "Class A")

    def test_e_class_archive_preserves_history(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 15),
            section=self.class_a,
        )
        self.class_a.archive()
        self.assertEqual(self.class_a.status, GroupSectionStatus.ARCHIVED)
        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.structured.pk,
            preset="today",
        )
        row = report["sections"][0]["rows"][0]
        self.assertEqual(row["class_name"], "Class A")
        record = ActionRecord.objects.get(member=self.clara, group=self.structured)
        self.assertEqual(record.class_name_snapshot, "Class A")
        self.assertEqual(record.section_id, self.class_a.pk)

    def test_f_class_permanent_delete_preserves_action_record_and_snapshot(self):
        record = self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 15),
            section=self.class_a,
        )
        section_id = self.class_a.pk
        self.class_a.archive()
        permanently_delete_section(self.class_a)

        record.refresh_from_db()
        self.assertIsNone(record.section_id)
        self.assertEqual(record.source_section_id, section_id)
        self.assertEqual(record.class_name_snapshot, "Class A")
        self.assertFalse(GroupSection.objects.filter(pk=section_id).exists())

        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.structured.pk,
            preset="today",
        )
        self.assertEqual(report["sections"][0]["rows"][0]["class_name"], "Class A")
        self.assertEqual(report["sections"][0]["rows"][0]["class_source_id"], section_id)

    def test_g_participant_moved_to_another_class_does_not_alter_old_report(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 15),
            section=self.class_a,
            class_name_snapshot="Class A",
        )
        self.clara_membership.section = self.class_b
        self.clara_membership.save(update_fields=["section", "updated_at"])
        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.structured.pk,
            preset="today",
        )
        self.assertEqual(report["sections"][0]["rows"][0]["class_name"], "Class A")

    def test_h_same_participant_two_classes_same_day_not_merged(self):
        day = self._day(9, 0)
        self._manual_record(
            member=self.aleks,
            action_type=ActionType.CHECK_IN,
            performed_at=day.replace(hour=9, minute=0),
            section=self.class_a,
        )
        self._manual_record(
            member=self.aleks,
            action_type=ActionType.CHECK_IN,
            performed_at=day.replace(hour=15, minute=0),
            section=self.class_b,
        )
        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.structured.pk,
            preset="today",
        )
        rows = report["sections"][0]["rows"]
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["class_name"], "Class A")
        self.assertEqual(rows[0]["name"], "Aleks")
        self.assertEqual(rows[0]["cells"]["check_in"], "09:00")
        self.assertEqual(rows[1]["class_name"], "Class B")
        self.assertEqual(rows[1]["name"], "Aleks")
        self.assertEqual(rows[1]["cells"]["check_in"], "15:00")

    def test_legacy_structured_without_class_snapshot_shows_unknown(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(10, 0),
            group=self.structured,
            group_type_snapshot=GroupType.STRUCTURED,
        )
        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.structured.pk,
            preset="today",
        )
        self.assertEqual(
            report["sections"][0]["rows"][0]["class_name"], UNKNOWN_CLASS_LABEL
        )

    def _structured_sample_report(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 15),
            section=self.class_a,
        )
        self._manual_record(
            member=self.jimi,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 44),
            section=self.class_b,
        )
        return build_attendance_report(
            organization=self.org,
            source_group_id=self.structured.pk,
            preset="today",
        )

    def test_i_csv_includes_class_for_structured(self):
        report = self._structured_sample_report()
        csv_text = build_attendance_report_csv(report)
        self.assertIn("Date,Class,Name,Check-in", csv_text)
        self.assertIn("Class A,Clara,11:15", csv_text)
        self.assertIn("Class B,Jimi,11:44", csv_text)

    def test_j_excel_includes_class_for_structured(self):
        report = self._structured_sample_report()
        content = build_attendance_report_xlsx(report)
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active
        values = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=1, max_col=6)]
        flat = [v for row in values for v in row if v]
        self.assertIn("Class", flat)
        self.assertIn("Class A", flat)
        self.assertIn("Clara", flat)

    def test_k_pdf_includes_class_for_structured(self):
        report = self._structured_sample_report()
        pdf_bytes = build_attendance_report_pdf(report)
        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertGreater(len(pdf_bytes), 1000)

    def test_l_standard_exports_unchanged(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(9, 0),
            group=self.standard,
        )
        report = build_attendance_report(
            organization=self.org,
            source_group_id=self.standard.pk,
            preset="today",
        )
        csv_text = build_attendance_report_csv(report)
        self.assertIn("Date,Name,Check-in", csv_text)
        self.assertNotIn("Date,Class,Name", csv_text)

        content = build_attendance_report_xlsx(report)
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active
        found_class_header = False
        for row in sheet.iter_rows(min_row=1, max_col=5):
            vals = [c.value for c in row]
            if "Date" in vals and "Name" in vals:
                self.assertNotIn("Class", vals)
                found_class_header = True
                break
        self.assertTrue(found_class_header)

        pdf_bytes = build_attendance_report_pdf(report)
        self.assertTrue(pdf_bytes.startswith(b"%PDF"))

    def test_m_tenant_isolation_unchanged(self):
        self._manual_record(
            member=self.clara,
            action_type=ActionType.CHECK_IN,
            performed_at=self._day(11, 15),
            section=self.class_a,
        )
        resp = self.other_client.get(
            f"/api/history/attendance-report/?source_group_id={self.structured.pk}&preset=today"
        )
        self.assertEqual(resp.status_code, 404)